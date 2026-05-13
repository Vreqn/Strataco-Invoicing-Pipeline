"""Per-manager / per-AP history XLS files.

Two flavours:

- Step 4 (manager pending-approval): per-manager XLS at
  `_state/toapprove_history/<YYYY-MM-DD>__<MANAGER_KEY>.xls`. Each day's run
  writes one file. The next day reads yesterday's file to compute old/new.

- Step 5 (AP approved baseline): rolling baseline at
  `_state/ap_approved_history/_latest__<AP_KEY>.xls`. Step 5 overwrites this
  every run after sending the notification, so the next run sees what was
  there last time and computes "new since last run".

Both files are simple XLS with one column ("Invoice") plus the ISO date.
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)


@dataclass
class OldNew:
    today: list[str]
    old: list[str]
    new: list[str]

    @property
    def total(self) -> int:
        return len(self.today)

    @property
    def old_count(self) -> int:
        return len(self.old)

    @property
    def new_count(self) -> int:
        return len(self.new)


def _read_invoice_column(xls_path: Path) -> list[str]:
    """Read one column of invoice filenames from a single-sheet XLS."""
    if not xls_path.exists():
        return []
    try:
        # openpyxl checks the filename extension before opening, and rejects
        # .xls names even though we're writing XLSX bytes into them (legacy
        # carryover from the N8n flow's filenames). Bypass the extension check
        # by handing it raw bytes; format autodetection on the buffer itself
        # then succeeds. See Codex P2 finding 2026-05-10.
        wb = load_workbook(
            filename=io.BytesIO(xls_path.read_bytes()),
            data_only=True, read_only=True,
        )
    except Exception as exc:
        logger.warning("could not open history xls %s: %s", xls_path, exc)
        return []
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h or "").strip().lower() for h in next(rows, [])]
    # Find a column called "invoice" / "filename" / "currentfilename" / "name"
    target_idx = None
    for i, h in enumerate(headers):
        if h in ("invoice", "filename", "currentfilename", "name", "invoicelist"):
            target_idx = i
            break
    out: list[str] = []
    for r in rows:
        if target_idx is not None and target_idx < len(r):
            v = r[target_idx]
        else:
            # Fallback: first non-empty cell
            v = next((c for c in r if c not in (None, "")), None)
        if v is None:
            continue
        s = str(v).strip()
        if s:
            out.append(s)
    return out


def _write_invoice_column(xls_path: Path, filenames: list[str], date_str: str) -> None:
    """Write a single-sheet XLS with columns Date + Invoice."""
    xls_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Invoice"])
    if not filenames:
        ws.append([date_str, ""])
    else:
        for fn in filenames:
            ws.append([date_str, fn])
    # Atomic save via tmp + replace
    tmp = xls_path.with_suffix(xls_path.suffix + ".tmp")
    wb.save(str(tmp))
    import os
    os.replace(tmp, xls_path)


def compute_old_new(today_files: list[str], previous_files: list[str]) -> OldNew:
    today = sorted({f for f in today_files if f}, key=str.lower)
    prev_set = {f.lower() for f in previous_files}
    old = [f for f in today if f.lower() in prev_set]
    new = [f for f in today if f.lower() not in prev_set]
    return OldNew(today=today, old=old, new=new)


# ----------------------------------------------------------------------
# Step 4 helpers (legacy — combined baseline)
# ----------------------------------------------------------------------


def read_yesterday_for_manager(yesterday_xls: Path) -> list[str]:
    return _read_invoice_column(yesterday_xls)


def write_today_for_manager(today_xls: Path, filenames: list[str], date_str: str) -> None:
    _write_invoice_column(today_xls, filenames, date_str)


# ----------------------------------------------------------------------
# Step 4 helpers (split — scanned vs notified)
#
# The old single-file design wrote today's snapshot unconditionally, then the
# next morning compared against it — so when an email failed to send, the
# unsent invoices silently aged from "new" into "old" and never got
# re-notified. The split adds a second file tracking only what was
# successfully emailed, and the diff reads from that one.
# ----------------------------------------------------------------------


def read_notified_for_manager(
    notified_xls: Path,
    legacy_xls: Path | None = None,
) -> list[str]:
    """Read the manager's "what was successfully emailed" baseline.

    Falls back to the pre-split combined file when the notified file doesn't
    exist yet (first run after the split lands). Pass the legacy path
    explicitly so this function stays decoupled from `paths.py`.
    """
    if notified_xls.exists():
        return _read_invoice_column(notified_xls)
    if legacy_xls is not None and legacy_xls.exists():
        return _read_invoice_column(legacy_xls)
    return []


def write_scanned_for_manager(
    scanned_xls: Path, filenames: list[str], date_str: str,
) -> None:
    """Diagnostic snapshot of what Step 4 saw on disk today. Always written."""
    _write_invoice_column(scanned_xls, filenames, date_str)


def write_notified_for_manager(
    notified_xls: Path, filenames: list[str], date_str: str,
) -> None:
    """Persist the list Step 4 successfully emailed. Only call after send succeeded."""
    _write_invoice_column(notified_xls, filenames, date_str)


# ----------------------------------------------------------------------
# Step 5 helpers (legacy — single rolling baseline)
# ----------------------------------------------------------------------


def read_ap_baseline(baseline_xls: Path) -> list[str]:
    return _read_invoice_column(baseline_xls)


def write_ap_baseline(baseline_xls: Path, filenames: list[str], date_str: str) -> None:
    _write_invoice_column(baseline_xls, filenames, date_str)


# ----------------------------------------------------------------------
# Step 5 helpers (split — scanned vs notified)
# ----------------------------------------------------------------------


def read_ap_notified_baseline(
    notified_xls: Path,
    legacy_xls: Path | None = None,
) -> list[str]:
    """Read the AP's "what was successfully emailed last run" baseline.

    Falls back to the pre-split rolling baseline when the notified file
    doesn't exist yet.
    """
    if notified_xls.exists():
        return _read_invoice_column(notified_xls)
    if legacy_xls is not None and legacy_xls.exists():
        return _read_invoice_column(legacy_xls)
    return []


def write_ap_scanned_baseline(
    scanned_xls: Path, filenames: list[str], date_str: str,
) -> None:
    """Diagnostic snapshot of what Step 5 saw in the AP folder. Always written."""
    _write_invoice_column(scanned_xls, filenames, date_str)


def write_ap_notified_baseline(
    notified_xls: Path, filenames: list[str], date_str: str,
) -> None:
    """Persist the list Step 5 successfully emailed. Only call after send succeeded."""
    _write_invoice_column(notified_xls, filenames, date_str)
