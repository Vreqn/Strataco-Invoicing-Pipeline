"""Read Strataplan_List.xlsx and build the lookup maps the steps need.

The XLS columns the N8n flow uses (preserved here):
- "Strata Plan"      e.g. "BCS 2707", "BCS-2707", "BCS2707"
- "Strata Name"
- "Address"
- "Strata Manager"   may be "First Last" or "First Last, Other Person"
- "Manager email"
- "AP Name"
- "AP email"
- "Status"           optional; row is active when missing or == 1

`load_plans()` returns a dict keyed by the *normalised* plan
(`BCS 2707` -> `BCS2707`) so the steps can look up by either filename or
subject token.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

from tools._lib import safe_io

_PLAN_NORM_RE = re.compile(r"[\s\-_./\\#№＃]+")


def _norm_plan(value: str | None) -> str:
    """Normalise a plan string: 'BCS 2707' / 'BCS-2707' -> 'BCS2707'.

    Also strips '#', '№', and fullwidth '＃' which appear occasionally.
    """
    if value is None:
        return ""
    s = str(value).upper().strip()
    s = _PLAN_NORM_RE.sub("", s)
    return s


def _primary_manager(value: str | None) -> str:
    """If multiple managers are listed (`A, B`), return the first."""
    if value is None:
        return ""
    return str(value).split(",")[0].strip()


def _split_emails(value: str | None) -> str:
    """N8n joined comma-separated emails with `;` for Graph; preserve that."""
    if value is None:
        return ""
    return str(value).replace(",", ";").strip()


def _key(name: str) -> str:
    """File-system safe upper-snake key for a person name."""
    s = (name or "").upper()
    s = re.sub(r"[^A-Z0-9]+", "_", s)
    s = re.sub(r"^_+|_+$", "", s)
    return s


@dataclass
class PlanRow:
    plan_norm: str
    plan_raw: str
    strata_name: str
    address: str
    manager_name: str
    manager_key: str
    manager_email: str
    ap_name: str
    ap_key: str
    ap_email: str
    status_active: bool


def _iter_rows(xlsx_path: Path) -> Iterator[dict]:
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h or "").strip() for h in next(rows)]
    for row in rows:
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        yield {h: v for h, v in zip(headers, row)}


def _validate_component(value: str, field: str, plan_raw: str) -> str:
    """Run sanitize_path_component on a row field; re-raise with row context."""
    if not value:
        return value
    try:
        return safe_io.sanitize_path_component(value)
    except ValueError as exc:
        raise ValueError(
            f"Strataplan_List.xlsx row for plan {plan_raw!r}: "
            f"{field} {value!r} is not a safe path component ({exc})"
        ) from exc


def _check_duplicate_routing(rows: list["PlanRow"]) -> None:
    """Fail loudly when two active rows share a plan_norm but disagree on routing.

    Pre-fix this would silently pick the first row (see `plan_to_manager`'s
    `setdefault`) and route invoices for the same plan to whichever manager
    happened to be listed first.
    """
    seen: dict[str, PlanRow] = {}
    for r in rows:
        if not r.status_active:
            continue
        prev = seen.get(r.plan_norm)
        if prev is None:
            seen[r.plan_norm] = r
            continue
        if prev.manager_name != r.manager_name or prev.ap_name != r.ap_name:
            raise ValueError(
                f"Duplicate plan rows with conflicting routing for "
                f"{r.plan_norm!r}: manager={prev.manager_name!r}/{r.manager_name!r}, "
                f"AP={prev.ap_name!r}/{r.ap_name!r}. Fix Strataplan_List.xlsx "
                f"before re-running."
            )


def load_plans(xlsx_path: Path) -> list[PlanRow]:
    """Load every active plan row from the XLS as `PlanRow` objects."""
    out: list[PlanRow] = []
    for r in _iter_rows(xlsx_path):
        plan_raw = str(r.get("Strata Plan") or "").strip()
        plan_norm = _norm_plan(plan_raw)
        if not plan_norm:
            continue

        manager_name = _validate_component(
            _primary_manager(r.get("Strata Manager")), "Strata Manager", plan_raw,
        )
        ap_name = _validate_component(
            str(r.get("AP Name") or "").strip(), "AP Name", plan_raw,
        )

        status_raw = r.get("Status")
        status_active = status_raw is None or str(status_raw).strip() in ("", "1")

        out.append(PlanRow(
            plan_norm=plan_norm,
            plan_raw=plan_raw,
            strata_name=str(r.get("Strata Name") or "").strip(),
            address=str(r.get("Address") or "").strip(),
            manager_name=manager_name,
            manager_key=_key(manager_name),
            manager_email=_split_emails(r.get("Manager email")),
            ap_name=ap_name,
            ap_key=_key(ap_name),
            ap_email=_split_emails(r.get("AP email")),
            status_active=status_active,
        ))
    _check_duplicate_routing(out)
    return out


def plan_to_manager(rows: list[PlanRow]) -> dict[str, PlanRow]:
    """Map plan_norm -> first PlanRow with a manager (active rows only)."""
    out: dict[str, PlanRow] = {}
    for r in rows:
        if not r.status_active or not r.manager_name:
            continue
        out.setdefault(r.plan_norm, r)
    return out


def plan_to_ap(rows: list[PlanRow]) -> dict[str, PlanRow]:
    """Map plan_norm -> first PlanRow with an AP (active rows only)."""
    out: dict[str, PlanRow] = {}
    for r in rows:
        if not r.status_active or not r.ap_name:
            continue
        out.setdefault(r.plan_norm, r)
    return out


def unique_managers(rows: list[PlanRow]) -> list[PlanRow]:
    """One PlanRow per manager (deduped by manager_key, active only)."""
    seen: set[str] = set()
    out: list[PlanRow] = []
    for r in rows:
        if not r.status_active or not r.manager_key:
            continue
        if r.manager_key in seen:
            continue
        seen.add(r.manager_key)
        out.append(r)
    return out


def unique_aps(rows: list[PlanRow]) -> list[PlanRow]:
    """One PlanRow per AP (deduped by ap_key, active only)."""
    seen: set[str] = set()
    out: list[PlanRow] = []
    for r in rows:
        if not r.status_active or not r.ap_key:
            continue
        if r.ap_key in seen:
            continue
        seen.add(r.ap_key)
        out.append(r)
    return out


def base_plan_index(rows: list[PlanRow]) -> dict[str, list[PlanRow]]:
    """For suffix-fallback in Step 5: 'LMS4193' -> [LMS4193C, LMS4193O, ...]

    Only includes rows where the plan ends in a single trailing letter.
    """
    out: dict[str, list[PlanRow]] = {}
    for r in rows:
        if not r.status_active:
            continue
        m = re.match(r"^(.*\d)([A-Z])$", r.plan_norm, re.IGNORECASE)
        if not m:
            continue
        base = m.group(1).upper()
        out.setdefault(base, []).append(r)
    return out
