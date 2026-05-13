"""Working-copy snapshot of Strataplan_List.xlsx.

The realistic baseline at 06:00 is that the local master XLS is open in Excel
on the automation machine — the operator leaves it open from the previous
day's editing session. A default `shutil.copy2` refresh would fail with
`PermissionError` on virtually every run.

This module reads the master via `CreateFileW` with `FILE_SHARE_READ |
FILE_SHARE_WRITE | FILE_SHARE_DELETE`. Excel itself opens xlsx files with
those same share flags, so Windows lets a matching-share read through even
while Excel holds the file. The bytes we read are whatever Excel last saved
to disk; unsaved in-memory edits are invisible, which matches the
"edit, save, close" protocol.

Step 1 owns the refresh. Steps 3-6 only call `require_fresh_snapshot()` and
halt the day if today's snapshot does not exist — stale routing is worse
than a delayed run, because re-routings (manager on vacation, new AP) must
take effect the day they are saved.
"""

from __future__ import annotations

import datetime as _dt
import os
import sys
from pathlib import Path

from openpyxl import load_workbook

from tools._lib import paths, safe_io


class SnapshotRefreshError(RuntimeError):
    """Raised when Step 1 cannot produce a fresh snapshot for today."""


class SnapshotStaleError(RuntimeError):
    """Raised when a step that requires today's snapshot finds none."""


_CHUNK = 1024 * 1024  # 1 MB


def _today_str() -> str:
    """America/Vancouver YYYY-MM-DD — same TZ used elsewhere in the pipeline."""
    try:
        from zoneinfo import ZoneInfo
        now = _dt.datetime.now(ZoneInfo("America/Vancouver"))
    except Exception:
        now = _dt.datetime.now()
    return now.strftime("%Y-%m-%d")


def _read_master_bytes(master: Path) -> bytes:
    """Read `master` even if Excel has it open (Windows share flags).

    On non-Windows platforms (dev/CI), falls back to a plain read — tests use
    that path. The share-flag behaviour is the load-bearing piece for prod.
    """
    if sys.platform != "win32":
        with open(master, "rb") as f:
            return f.read()

    import ctypes
    from ctypes import wintypes

    GENERIC_READ = 0x80000000
    FILE_SHARE_READ = 0x00000001
    FILE_SHARE_WRITE = 0x00000002
    FILE_SHARE_DELETE = 0x00000004
    OPEN_EXISTING = 3
    FILE_ATTRIBUTE_NORMAL = 0x80
    INVALID_HANDLE_VALUE = wintypes.HANDLE(-1).value

    # use_last_error=True wires ctypes.get_last_error() to the Win32
    # GetLastError state, so a sharing-violation or access-denied code
    # actually surfaces in the SnapshotRefreshError message.
    kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)

    CreateFileW = kernel32.CreateFileW
    CreateFileW.argtypes = [
        wintypes.LPCWSTR,
        wintypes.DWORD,
        wintypes.DWORD,
        ctypes.c_void_p,
        wintypes.DWORD,
        wintypes.DWORD,
        wintypes.HANDLE,
    ]
    CreateFileW.restype = wintypes.HANDLE

    ReadFile = kernel32.ReadFile
    ReadFile.argtypes = [
        wintypes.HANDLE,
        ctypes.c_void_p,
        wintypes.DWORD,
        ctypes.POINTER(wintypes.DWORD),
        ctypes.c_void_p,
    ]
    ReadFile.restype = wintypes.BOOL

    CloseHandle = kernel32.CloseHandle
    CloseHandle.argtypes = [wintypes.HANDLE]
    CloseHandle.restype = wintypes.BOOL

    handle = CreateFileW(
        str(master),
        GENERIC_READ,
        FILE_SHARE_READ | FILE_SHARE_WRITE | FILE_SHARE_DELETE,
        None,
        OPEN_EXISTING,
        FILE_ATTRIBUTE_NORMAL,
        None,
    )
    if handle == INVALID_HANDLE_VALUE or handle is None:
        err = ctypes.get_last_error() or ctypes.GetLastError()
        raise SnapshotRefreshError(
            f"CreateFileW failed for {master} (Win32 error {err})"
        )

    try:
        chunks: list[bytes] = []
        buf = (ctypes.c_ubyte * _CHUNK)()
        bytes_read = wintypes.DWORD(0)
        while True:
            ok = ReadFile(handle, buf, _CHUNK, ctypes.byref(bytes_read), None)
            if not ok:
                err = ctypes.GetLastError()
                raise SnapshotRefreshError(
                    f"ReadFile failed for {master} (Win32 error {err})"
                )
            n = bytes_read.value
            if n == 0:
                break
            chunks.append(bytes(buf[:n]))
        return b"".join(chunks)
    finally:
        CloseHandle(handle)


def _verify_xlsx(path: Path) -> None:
    """Open the file to confirm the bytes are valid XLSX.

    Called against the staged tmp file *before* it is promoted to the
    snapshot path — so a corrupt read can never replace a previously-good
    snapshot or be authorised by a leftover today-marker.
    """
    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        try:
            # Touch the active sheet so openpyxl actually parses, not just opens.
            _ = wb.active.max_row
        finally:
            wb.close()
    except Exception as exc:
        raise SnapshotRefreshError(
            f"file at {path} did not validate as XLSX: {exc}"
        ) from exc


def refresh_snapshot() -> Path:
    """Stage master -> tmp, validate the tmp, then atomic-rename + write marker.

    Verify-before-publish: the snapshot file at the published path is only
    replaced after the new bytes parse as XLSX. The marker is rewritten last,
    so a failure at any earlier step leaves the marker stale (or absent) and
    Steps 3-6 halt rather than read corrupt bytes.

    Returns the snapshot path on success. Raises `SnapshotRefreshError` on
    any failure.
    """
    master = paths.strataplan_xlsx()
    if not master.exists():
        raise SnapshotRefreshError(f"master file not found: {master}")

    snapshot = paths.strataplan_snapshot_xlsx()
    marker = paths.strataplan_snapshot_marker()

    try:
        data = _read_master_bytes(master)
    except SnapshotRefreshError:
        raise
    except Exception as exc:
        raise SnapshotRefreshError(
            f"failed to read master {master}: {exc}"
        ) from exc

    if not data:
        raise SnapshotRefreshError(f"master {master} read as 0 bytes")

    # Stage the new bytes to a tmp path next to the final snapshot, validate
    # there, only then promote. The tmp filename must still end in `.xlsx`
    # because openpyxl rejects unknown extensions on load.
    safe_io.ensure_parent(snapshot)
    tmp = snapshot.parent / f"{snapshot.stem}.tmp.{os.getpid()}{snapshot.suffix}"
    try:
        try:
            with open(tmp, "wb") as f:
                f.write(data)
                f.flush()
                os.fsync(f.fileno())
        except Exception as exc:
            raise SnapshotRefreshError(
                f"failed to stage snapshot at {tmp}: {exc}"
            ) from exc

        _verify_xlsx(tmp)

        try:
            os.replace(tmp, snapshot)
        except Exception as exc:
            raise SnapshotRefreshError(
                f"failed to promote staged snapshot {tmp} -> {snapshot}: {exc}"
            ) from exc
    finally:
        # Best-effort cleanup if anything above raised before the rename
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass

    today = _today_str()
    try:
        safe_io.atomic_write_bytes(marker, today.encode("utf-8"))
    except Exception as exc:
        raise SnapshotRefreshError(
            f"failed to write marker {marker}: {exc}"
        ) from exc

    return snapshot


def require_fresh_snapshot() -> Path:
    """Assert the marker contains today's date; return the snapshot path.

    Used by Steps 3-6 so they refuse to run on a day where Step 1 did not
    successfully refresh.
    """
    marker = paths.strataplan_snapshot_marker()
    snapshot = paths.strataplan_snapshot_xlsx()

    if not marker.exists():
        raise SnapshotStaleError(
            f"no snapshot marker at {marker} — Step 1 has not run today"
        )
    if not snapshot.exists():
        raise SnapshotStaleError(
            f"marker present but snapshot missing at {snapshot}"
        )

    try:
        marker_date = marker.read_text(encoding="utf-8").strip()
    except Exception as exc:
        raise SnapshotStaleError(
            f"could not read marker {marker}: {exc}"
        ) from exc

    today = _today_str()
    if marker_date != today:
        raise SnapshotStaleError(
            f"snapshot marker is {marker_date!r}, expected {today!r}"
        )

    return snapshot
